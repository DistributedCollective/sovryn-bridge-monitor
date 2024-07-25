from dataclasses import dataclass
from decimal import Decimal
from datetime import datetime
from logging import getLogger
from tempfile import NamedTemporaryFile

from pyramid.view import view_config
from pyramid.response import Response
from sqlalchemy.orm import Session
from sqlalchemy.sql import select
from sqlalchemy.sql.expression import func
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font

from .pnl import _parse_time_range
from bridge_monitor.models.ledger_entry import LedgerEntry, LedgerAccount
from ..models.ledger_meta import LedgerUpdateMeta

logger = getLogger(__name__)


@dataclass
class AccountBalance:
    account_id: int
    account_name: str
    balance: Decimal
    credit: Decimal
    debit: Decimal
    is_debit: bool


@dataclass
class EntryDisplay:
    account_name: str
    value: Decimal
    timestamp: datetime
    tx_hash: str


@view_config(
    route_name="ledger",
    renderer="bridge_monitor:templates/ledger.jinja2",
)
def ledger(request):
    def get_account_name(account_id: int, account_list: list[AccountBalance]) -> str:
        for acc in account_list:
            if acc.account_id == abs(account_id):
                return acc.account_name
        return f"Unknown account {account_id}"

    dbsession: Session = request.dbsession

    start, end, errors, time_filter = _parse_time_range(request)

    amount_in_page = 100
    first_entry_number = request.params.get("first_entry", 1)
    first_entry_number = max(1, int(first_entry_number))
    last_entry_number = first_entry_number + amount_in_page - 1

    accounts = (
        dbsession.execute(
            select(LedgerAccount).where(LedgerAccount.id > 0).order_by(LedgerAccount.id)
        )
        .scalars()
        .all()
    )

    ledger_last_updated_at = dbsession.execute(
        select(LedgerUpdateMeta.timestamp).order_by(LedgerUpdateMeta.timestamp.desc())
    ).scalar()

    account_balances = []
    ledger_entries = []

    if start and end:
        for account in accounts:
            credit = dbsession.execute(
                select(func.sum(LedgerEntry.value)).where(
                    LedgerEntry.account_id == -account.id,
                    LedgerEntry.timestamp >= start,
                    LedgerEntry.timestamp <= end,
                )
            ).scalar()
            credit = Decimal(0) if credit is None else abs(credit)

            debit = dbsession.execute(
                select(func.sum(LedgerEntry.value)).where(
                    LedgerEntry.account_id == account.id,
                    LedgerEntry.timestamp >= start,
                    LedgerEntry.timestamp <= end,
                )
            ).scalar()
            debit = Decimal(0) if debit is None else debit

            balance = (debit - credit) * (1 if account.is_debit else -1)

            account_balances.append(
                AccountBalance(
                    account.id, account.name, balance, credit, debit, account.is_debit
                )
            )

        ledger_entries = (
            dbsession.execute(
                select(LedgerEntry)
                .where(LedgerEntry.timestamp >= start, LedgerEntry.timestamp <= end)
                .order_by(LedgerEntry.timestamp, LedgerEntry.id)
            )
            .scalars()
            .all()
        )

        # if request type is post then download all the data in excel format
        if request.method == "POST":
            wb = Workbook()
            curr_sheet = wb.active

            bold_style = NamedStyle(name="bold")
            bold_style.font = Font(bold=True)
            number_style = NamedStyle(name="number")
            number_style.number_format = "0.000000000000"

            curr_sheet.title = "Accounts"
            curr_sheet.append((f"{start} - {end}",))
            curr_sheet.cell(row=1, column=1).style = bold_style
            sub_headings = [
                "Account Name",
                "Credit",
                "Debit",
                "Account type",
                "Balance delta",
            ]
            for i, heading in enumerate(sub_headings, start=1):
                curr_sheet.cell(row=2, column=i).value = heading
                curr_sheet.cell(row=2, column=i).style = bold_style
            for j, account in enumerate(account_balances, start=3):
                curr_sheet.append(
                    [
                        account.account_name,
                        account.credit,
                        account.debit,
                        "debit" if account.is_debit else "credit",
                        account.balance,
                    ]
                )
                for i in range(len(sub_headings)):
                    curr_sheet.cell(row=j, column=i + 1).style = number_style

            curr_sheet["H2"] = "Total change in credit accounts"
            curr_sheet["H2"].style = bold_style
            curr_sheet["H3"] = "Total change in debit accounts"
            curr_sheet["H3"].style = bold_style
            curr_sheet["I2"] = sum(
                [
                    account.balance
                    for account in account_balances
                    if not account.is_debit
                ]
            )
            curr_sheet["I3"] = sum(
                [account.balance for account in account_balances if account.is_debit]
            )

            curr_sheet = wb.create_sheet(title="Ledger Entries")
            ledger_sub_headings = ["Account Name", "Value", "Timestamp", "Tx Hash"]
            for i, heading in enumerate(ledger_sub_headings, start=1):
                curr_sheet.cell(row=1, column=i).value = heading
                curr_sheet.cell(row=1, column=i).style = bold_style

            for row, entry in enumerate(ledger_entries, start=2):
                curr_sheet.append(
                    [
                        get_account_name(entry.account_id, account_balances),
                        entry.value,
                        entry.timestamp.replace(tzinfo=None),
                        entry.tx_hash,
                    ]
                )
                curr_sheet.cell(row=row, column=2).style = number_style
            response = Response(content_type="application/vnd.ms-excel")
            response.content_disposition = "attachment;filename=ledger.xlsx"
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                response.body = tmp.read()
                return response

        ledger_entries = ledger_entries[first_entry_number - 1 : last_entry_number]
        ledger_entries = [
            EntryDisplay(
                get_account_name(entry.account_id, account_balances),
                entry.value,
                entry.timestamp,
                entry.tx_hash,
            )
            for entry in ledger_entries
        ]

        last_entry_number = len(ledger_entries) + first_entry_number - 1
        if first_entry_number > last_entry_number:
            first_entry_number -= amount_in_page
    return {
        "first_entry": first_entry_number,
        "last_entry": last_entry_number,
        "amount_in_page": amount_in_page,
        "accounts": accounts,
        "account_balances": account_balances,
        "ledger_entries": ledger_entries,
        "ledger_last_updated_at": ledger_last_updated_at,
    }
